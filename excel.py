from openpyxl import load_workbook, Workbook

data1 = load_workbook('章节考勤数据.xlsx')
print(data1.sheetnames)
table1 = data1.get_sheet_by_name('考勤人数')
data2 = load_workbook('python.xlsx')
print(data2.sheetnames)
table2 = data2.get_sheet_by_name('Sheet1')


rs = [r for r in table1]
# print(len(set([r[4].value for r in rs])))
rl = [i for i in table2]

wb = Workbook()
sheet1 = wb.active
sheet1.title = '18班信息'

sheet1 = wb.create_sheet()

for i in rs:
    print(i[0].value)
    for x in rl:
        print(x[7].value)
        if str(i[0].value).lower() == str(x[7].value).lower():
            # print(i[3].value)
            sheet1.append([oo[2].value for oo in x])
wb.save('18项目信息.xlsx')

# table2 = data.get_sheet_by_name('qq人数')
#
# rs = [r for r in table1]
# rl = [i for i in table2]
# # print(rl)
#
# wb = Workbook()
# sheet3 = wb.active
# sheet3.title = '到人数'
# sheet3 = wb.create_sheet()
#
# for i in rs:
#     # print(i)
#     for x in rl:
#         if str(i[3].value) == str(x[2].value):
#             print(i[3].value)
#             sheet3.append([oo.value for oo in i])
#
#
# wb.save('new.xlsx')
#
# # print(rs[5][3].value)
# # print('*'*80)
# # qq1 = [r[3].value for r in rs]
# # qq2 = [str(r[4].value) for r in rl]
# # print(qq1)
# # print(qq2)
# # qq1 = set(qq1)
# # qq2 = set(qq2)
# #
# # print(len(qq1))
# # print(len(qq2))
# # print(qq1)
# # print(qq2)
# # qq3 = qq1&qq2
# # print(qq3)
# # print(len(qq3))
#
#
